# =========================================================
# Excel Report Routes
# Autonomous AI DBA Operations Platform
# =========================================================

import os
import smtplib
from datetime import datetime
from email.message import EmailMessage
from typing import Any, Dict, List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Request
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from app.reporting.monthly_excel_report_generator import (
    generate_monthly_dba_excel_report,
    get_monthly_report_file_path
)


# =========================================================
# ROUTER AND TEMPLATE CONFIGURATION
# =========================================================

router = APIRouter()

templates = Jinja2Templates(
    directory="templates"
)


# =========================================================
# CONFIGURATION
# =========================================================

EXCEL_REPORT_FOLDER = "excel_reports"

ALLOWED_EXCEL_EXTENSIONS = [
    ".xlsx"
]

MAX_PREVIEW_ROWS = 100

MAX_PREVIEW_COLUMNS = 25


# =========================================================
# SAFE EXCEL REPORT PATH
# =========================================================

def resolve_safe_excel_report_path(
    report_path: str
) -> Optional[str]:
    """
    Resolve and validate Excel report path.
    """

    if not report_path:

        return None

    normalized_path = str(
        report_path
    ).replace(
        "\\",
        "/"
    ).strip()

    file_extension = os.path.splitext(
        normalized_path
    )[1].lower()

    if file_extension not in ALLOWED_EXCEL_EXTENSIONS:

        return None

    if not normalized_path.startswith(
        f"{EXCEL_REPORT_FOLDER}/"
    ):

        return None

    absolute_path = os.path.abspath(
        normalized_path
    )

    project_root = os.path.abspath(
        "."
    )

    if not absolute_path.startswith(
        project_root
    ):

        return None

    if not os.path.exists(
        absolute_path
    ):

        return None

    if not os.path.isfile(
        absolute_path
    ):

        return None

    return absolute_path


# =========================================================
# EXCEL REPORT TYPE
# =========================================================

def get_excel_report_type(
    file_name: str
) -> str:
    """
    Identify Excel report type.
    """

    lower_name = file_name.lower()

    if "monthly" in lower_name:

        return "Monthly DBA Operations Report"

    if "health" in lower_name:

        return "Excel Health Analytics Report"

    if "approval" in lower_name:

        return "Excel Approval Report"

    if "governance" in lower_name:

        return "Excel Governance Report"

    if "performance" in lower_name:

        return "Excel Performance Report"

    return "Excel Report"


# =========================================================
# LOAD EXCEL REPORTS
# =========================================================

def load_excel_reports() -> Dict[str, Any]:
    """
    Load only Excel reports from excel_reports folder.
    """

    excel_reports: List[Dict[str, Any]] = []

    if not os.path.exists(
        EXCEL_REPORT_FOLDER
    ):

        return {
            "total_reports": 0,
            "latest_report": "-",
            "reports": []
        }

    for root, directories, files in os.walk(
        EXCEL_REPORT_FOLDER
    ):

        for file_name in files:

            if not file_name.endswith(
                ".xlsx"
            ):

                continue

            file_path = os.path.join(
                root,
                file_name
            ).replace(
                "\\",
                "/"
            )

            try:

                modified_timestamp = os.path.getmtime(
                    file_path
                )

                modified_at = datetime.fromtimestamp(
                    modified_timestamp
                ).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )

                file_size_kb = round(
                    os.path.getsize(
                        file_path
                    ) / 1024,
                    2
                )

            except Exception:

                modified_timestamp = 0
                modified_at = "-"
                file_size_kb = 0

            encoded_file_path = quote(
                file_path
            )

            excel_reports.append(
                {
                    "report_name": file_name,
                    "report_type": get_excel_report_type(
                        file_name
                    ),
                    "status": "Available",
                    "file_path": file_path,
                    "file_size_kb": file_size_kb,
                    "modified_at": modified_at,
                    "modified_timestamp": modified_timestamp,
                    "view_url": f"/reports/view?file_path={encoded_file_path}",
                    "download_url": f"/reports/download?file_path={encoded_file_path}"
                }
            )

    excel_reports.sort(
        key=lambda item: item.get(
            "modified_timestamp",
            0
        ),
        reverse=True
    )

    latest_report = "-"

    if excel_reports:

        latest_report = excel_reports[
            0
        ].get(
            "report_name",
            "-"
        )

    return {
        "total_reports": len(
            excel_reports
        ),
        "latest_report": latest_report,
        "reports": excel_reports
    }


# =========================================================
# EXCEL PREVIEW READER
# =========================================================

def read_excel_workbook_preview(
    file_path: str
) -> Dict[str, Any]:
    """
    Read Excel workbook content for browser preview.
    """

    try:

        from openpyxl import load_workbook

    except Exception as error:

        return {
            "status": "FAILED",
            "error": (
                "openpyxl package is required to preview Excel reports. "
                f"Error: {str(error)}"
            ),
            "sheets": []
        }

    try:

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True
        )

        sheet_previews = []

        for worksheet in workbook.worksheets:

            rows = []
            row_count = 0

            for row in worksheet.iter_rows(
                values_only=True
            ):

                row_count += 1

                if row_count > MAX_PREVIEW_ROWS:

                    break

                normalized_row = []

                for cell_value in list(
                    row
                )[:MAX_PREVIEW_COLUMNS]:

                    if cell_value is None:

                        normalized_row.append(
                            ""
                        )

                    else:

                        normalized_row.append(
                            str(
                                cell_value
                            )
                        )

                rows.append(
                    normalized_row
                )

            sheet_previews.append(
                {
                    "sheet_name": worksheet.title,
                    "rows": rows,
                    "preview_rows": len(
                        rows
                    )
                }
            )

        workbook.close()

        return {
            "status": "SUCCESS",
            "error": None,
            "sheets": sheet_previews
        }

    except Exception as error:

        return {
            "status": "FAILED",
            "error": f"Unable to preview Excel report. Error: {str(error)}",
            "sheets": []
        }


# =========================================================
# EMAIL LATEST REPORT
# =========================================================

def email_latest_report() -> Dict[str, str]:
    """
    Email latest monthly report using SMTP environment variables.
    """

    report_path = get_monthly_report_file_path()

    if not os.path.exists(
        report_path
    ):

        return {
            "status": "FAILED",
            "message": "Monthly Excel report does not exist. Generate report first."
        }

    smtp_host = os.getenv(
        "SMTP_HOST"
    )

    smtp_port = os.getenv(
        "SMTP_PORT",
        "587"
    )

    smtp_username = os.getenv(
        "SMTP_USERNAME"
    )

    smtp_password = os.getenv(
        "SMTP_PASSWORD"
    )

    email_from = os.getenv(
        "EMAIL_FROM"
    )

    email_to = os.getenv(
        "EMAIL_TO"
    )

    if not all(
        [
            smtp_host,
            smtp_username,
            smtp_password,
            email_from,
            email_to
        ]
    ):

        return {
            "status": "CONFIGURATION_PENDING",
            "message": "SMTP configuration is missing in .env file."
        }

    try:

        message = EmailMessage()

        message[
            "Subject"
        ] = "DBA Monthly Operations Report"

        message[
            "From"
        ] = email_from

        message[
            "To"
        ] = email_to

        message.set_content(
            "Please find attached the DBA Monthly Operations Report."
        )

        with open(
            report_path,
            "rb"
        ) as file:

            message.add_attachment(
                file.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(
                    report_path
                )
            )

        with smtplib.SMTP(
            smtp_host,
            int(
                smtp_port
            )
        ) as server:

            server.starttls()

            server.login(
                smtp_username,
                smtp_password
            )

            server.send_message(
                message
            )

        return {
            "status": "SENT",
            "message": "Monthly Excel report emailed successfully."
        }

    except Exception as error:

        return {
            "status": "FAILED",
            "message": str(
                error
            )
        }


# =========================================================
# REPORTS CENTER
# =========================================================

@router.get("/reports")
def reports(
    request: Request,
    message: Optional[str] = None,
    status: Optional[str] = None
):
    """
    Render Excel Reports Center.
    """

    reports_data = load_excel_reports()

    return templates.TemplateResponse(
        request=request,
        name="reports.html",
        context={
            "reports": reports_data,
            "reports_data": reports_data,
            "report_files": reports_data.get(
                "reports",
                []
            ),
            "total_reports": reports_data.get(
                "total_reports",
                0
            ),
            "latest_report": reports_data.get(
                "latest_report",
                "-"
            ),
            "message": message,
            "message_status": status
        }
    )


# =========================================================
# GENERATE MONTHLY REPORT
# =========================================================

@router.post("/reports/generate/monthly")
def generate_monthly_report():
    """
    Generate monthly Excel report.
    """

    result = generate_monthly_dba_excel_report(
        trigger_source="Reports UI - Monthly Report"
    )

    return RedirectResponse(
        url=f"/reports?status={result.get('overall_status')}&message={result.get('message')}",
        status_code=303
    )


# =========================================================
# GENERATE WEEKLY REPORT
# =========================================================

@router.post("/reports/generate/weekly")
def generate_weekly_report():
    """
    Generate weekly snapshot into monthly Excel workbook.
    """

    result = generate_monthly_dba_excel_report(
        trigger_source="Reports UI - Weekly Snapshot",
        notes="Weekly snapshot added into monthly Excel workbook."
    )

    return RedirectResponse(
        url=f"/reports?status={result.get('overall_status')}&message=Weekly snapshot updated in monthly Excel workbook.",
        status_code=303
    )


# =========================================================
# DOWNLOAD LATEST REPORT
# =========================================================

@router.get("/reports/download/latest")
def download_latest_report():
    """
    Download current monthly report.
    """

    report_path = get_monthly_report_file_path()

    safe_path = resolve_safe_excel_report_path(
        report_path.replace(
            "\\",
            "/"
        )
    )

    if not safe_path:

        return HTMLResponse(
            content="""
            <html>
                <body style="font-family: Arial; padding: 30px;">
                    <h3>Monthly Excel report not found.</h3>
                    <p>Please generate the monthly report first.</p>
                    <a href="/reports">Back to Reports</a>
                </body>
            </html>
            """,
            status_code=404
        )

    return FileResponse(
        path=safe_path,
        filename=os.path.basename(
            safe_path
        ),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================================
# EMAIL LATEST REPORT
# =========================================================

@router.post("/reports/email/latest")
def email_latest_report_route():
    """
    Email latest monthly report.
    """

    result = email_latest_report()

    return RedirectResponse(
        url=f"/reports?status={result.get('status')}&message={result.get('message')}",
        status_code=303
    )


# =========================================================
# VIEW REPORT
# =========================================================

@router.get("/reports/view")
def view_report(
    request: Request,
    file_path: str
):
    """
    View Excel report content in browser.
    """

    safe_path = resolve_safe_excel_report_path(
        file_path
    )

    if not safe_path:

        return HTMLResponse(
            content="""
            <html>
                <body style="font-family: Arial; padding: 30px;">
                    <h3>Invalid or missing Excel report file.</h3>
                    <p>The selected Excel report file could not be found or is not allowed.</p>
                    <a href="/reports">Back to Reports</a>
                </body>
            </html>
            """,
            status_code=404
        )

    report_name = os.path.basename(
        safe_path
    )

    workbook_preview = read_excel_workbook_preview(
        safe_path
    )

    return templates.TemplateResponse(
        request=request,
        name="report_view.html",
        context={
            "report_name": report_name,
            "report_path": file_path,
            "file_extension": ".xlsx",
            "preview_status": workbook_preview.get(
                "status",
                "FAILED"
            ),
            "preview_error": workbook_preview.get(
                "error"
            ),
            "sheets": workbook_preview.get(
                "sheets",
                []
            )
        }
    )


# =========================================================
# DOWNLOAD REPORT
# =========================================================

@router.get("/reports/download")
def download_report(
    file_path: str
):
    """
    Download selected Excel report.
    """

    safe_path = resolve_safe_excel_report_path(
        file_path
    )

    if not safe_path:

        return HTMLResponse(
            content="""
            <html>
                <body style="font-family: Arial; padding: 30px;">
                    <h3>Invalid or missing Excel report file.</h3>
                    <p>The selected Excel report file could not be found or is not allowed.</p>
                    <a href="/reports">Back to Reports</a>
                </body>
            </html>
            """,
            status_code=404
        )

    return FileResponse(
        path=safe_path,
        filename=os.path.basename(
            safe_path
        ),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )