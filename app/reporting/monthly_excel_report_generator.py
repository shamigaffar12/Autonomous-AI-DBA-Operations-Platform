# =========================================================
# Monthly Excel Report Generator
# Autonomous AI DBA Operations Platform
# =========================================================

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# FOLDER CONFIGURATION
# =========================================================

EXCEL_REPORT_FOLDER = "excel_reports"

REPOSITORY_FOLDER = "repository"

APPROVAL_REQUEST_FOLDER = "approval_requests"

AUDIT_LOG_FOLDER = "audit_logs"


# =========================================================
# SOURCE FILES
# =========================================================

INCIDENTS_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "incidents.json"
)

ACTIONS_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "actions.json"
)

PENDING_APPROVALS_FILE = os.path.join(
    APPROVAL_REQUEST_FOLDER,
    "pending_approvals.json"
)

APPROVAL_HISTORY_FILE = os.path.join(
    APPROVAL_REQUEST_FOLDER,
    "approval_history.json"
)

EXECUTION_HISTORY_FILE = os.path.join(
    APPROVAL_REQUEST_FOLDER,
    "execution_history.json"
)

GOVERNANCE_AUDIT_FILE = os.path.join(
    APPROVAL_REQUEST_FOLDER,
    "governance_audit_log.json"
)

PERFORMANCE_METRICS_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "performance_metrics.json"
)

NLP_ACTIONS_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "nlp_actions.json"
)

BACKUP_MONITORING_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "backup_monitoring.json"
)

WORKFLOW_HISTORY_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "workflow_history.json"
)

AI_RECOMMENDATIONS_FILE = os.path.join(
    REPOSITORY_FOLDER,
    "ai_recommendations.json"
)


# =========================================================
# SHEET CONFIGURATION
# =========================================================

SHEET_DEFINITIONS = {
    "Summary": [
        "Date",
        "Total Incidents",
        "Healthy Status Count",
        "Attention Status Count",
        "Actions Executed",
        "Approvals Created",
        "Approvals Approved",
        "Approvals Rejected"
    ],
    "Daily Incidents": [
        "Timestamp",
        "Incident ID",
        "Incident Type",
        "Severity",
        "Status",
        "Risk Level",
        "Database Name",
        "Details",
        "AI Recommendation"
    ],
    "Performance Metrics": [
        "Timestamp",
        "CPU %",
        "Memory %",
        "Blocking Sessions",
        "Deadlocks",
        "Long Running Queries",
        "Database Size (GB)",
        "Active Connections"
    ],
    "NLP Actions": [
        "Timestamp",
        "User",
        "Natural Language Query",
        "Generated SQL",
        "Risk Classification",
        "Approval Status",
        "Execution Status"
    ],
    "Approvals Audit": [
        "Approval ID",
        "Timestamp",
        "Requestor",
        "Action Type",
        "Risk Level",
        "Approval Status",
        "Approver",
        "Approval Time"
    ],
    "Execution History": [
        "Execution ID",
        "Timestamp",
        "Action Executed",
        "Database",
        "Execution Status",
        "Duration",
        "Output Summary"
    ],
    "Governance Audit": [
        "Timestamp",
        "User",
        "Activity Type",
        "Object Affected",
        "Result",
        "Comments"
    ],
    "AI Recommendations": [
        "Timestamp",
        "Incident ID",
        "Root Cause",
        "Recommendation Category",
        "Recommendation Details",
        "Confidence Score"
    ],
    "Backup Monitoring": [
        "Timestamp",
        "Database",
        "Last Backup Time",
        "Backup Type",
        "Backup Status",
        "Recovery Model"
    ],
    "Agentic Workflow History": [
        "Timestamp",
        "Workflow Name",
        "Trigger Source",
        "Status",
        "Duration",
        "Result Summary"
    ]
}


# =========================================================
# COMMON UTILITIES
# =========================================================

def ensure_excel_folder() -> None:
    """
    Ensure Excel output folder exists.
    """

    os.makedirs(
        EXCEL_REPORT_FOLDER,
        exist_ok=True
    )


def current_timestamp() -> str:
    """
    Return current timestamp.
    """

    return datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )


def current_date() -> str:
    """
    Return current date.
    """

    return datetime.now().strftime(
        "%Y-%m-%d"
    )


def get_monthly_report_file_path() -> str:
    """
    Return monthly Excel report path.

    Format:
    DBA_Monthly_Report_2026_06.xlsx
    """

    ensure_excel_folder()

    now = datetime.now()

    file_name = f"DBA_Monthly_Report_{now.year}_{now.month:02d}.xlsx"

    return os.path.join(
        EXCEL_REPORT_FOLDER,
        file_name
    )


def safe_value(
    value: Any
) -> str:
    """
    Convert any value to safe Excel string value.
    """

    if value is None:

        return ""

    if isinstance(
        value,
        (
            dict,
            list
        )
    ):

        try:

            return json.dumps(
                value,
                ensure_ascii=False
            )

        except Exception:

            return str(
                value
            )

    return str(
        value
    )


def load_json_file(
    file_path: str,
    default_value: Optional[Any] = None
) -> Any:
    """
    Load JSON file safely.
    """

    if default_value is None:

        default_value = []

    try:

        if not os.path.exists(
            file_path
        ):

            return default_value

        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:

            return json.load(
                file
            )

    except Exception:

        return default_value


def ensure_list(
    value: Any
) -> List[Dict[str, Any]]:
    """
    Ensure input is list.
    """

    if isinstance(
        value,
        list
    ):

        return value

    return []


def get_first_available(
    data: Dict[str, Any],
    keys: List[str],
    default_value: Any = ""
) -> Any:
    """
    Return first available value from dictionary.
    """

    for key in keys:

        if key in data and data.get(
            key
        ) not in [
            None,
            ""
        ]:

            return data.get(
                key
            )

    return default_value


def normalize_timestamp(
    value: Any
) -> str:
    """
    Normalize timestamp value.
    """

    if value:

        return safe_value(
            value
        ).replace(
            "T",
            " "
        ).split(
            "."
        )[0]

    return current_timestamp()


# =========================================================
# EXISTING WORKBOOK DATA
# =========================================================

def read_existing_rows(
    file_path: str,
    sheet_name: str,
    headers: List[str]
) -> List[Dict[str, Any]]:
    """
    Read existing sheet rows from monthly workbook.
    """

    if not os.path.exists(
        file_path
    ):

        return []

    try:

        workbook = load_workbook(
            file_path
        )

        if sheet_name not in workbook.sheetnames:

            workbook.close()

            return []

        worksheet = workbook[
            sheet_name
        ]

        rows = []

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if not row:

                continue

            if all(
                cell is None or str(
                    cell
                ).strip() == ""
                for cell in row
            ):

                continue

            record = {}

            for index, header in enumerate(
                headers
            ):

                value = ""

                if index < len(
                    row
                ):

                    value = row[
                        index
                    ]

                record[
                    header
                ] = safe_value(
                    value
                )

            rows.append(
                record
            )

        workbook.close()

        return rows

    except Exception:

        return []


def unique_rows(
    rows: List[Dict[str, Any]],
    key_fields: List[str]
) -> List[Dict[str, Any]]:
    """
    Remove duplicate rows using key fields.
    """

    result = []
    seen = set()

    for row in rows:

        key = tuple(
            safe_value(
                row.get(
                    field,
                    ""
                )
            )
            for field in key_fields
        )

        if key in seen:

            continue

        seen.add(
            key
        )

        result.append(
            row
        )

    return result


# =========================================================
# DATA LOADERS
# =========================================================

def load_platform_data() -> Dict[str, List[Dict[str, Any]]]:
    """
    Load all reporting source data from repositories.
    """

    return {
        "incidents": ensure_list(
            load_json_file(
                INCIDENTS_FILE,
                []
            )
        ),
        "actions": ensure_list(
            load_json_file(
                ACTIONS_FILE,
                []
            )
        ),
        "pending_approvals": ensure_list(
            load_json_file(
                PENDING_APPROVALS_FILE,
                []
            )
        ),
        "approval_history": ensure_list(
            load_json_file(
                APPROVAL_HISTORY_FILE,
                []
            )
        ),
        "execution_history": ensure_list(
            load_json_file(
                EXECUTION_HISTORY_FILE,
                []
            )
        ),
        "governance_audit": ensure_list(
            load_json_file(
                GOVERNANCE_AUDIT_FILE,
                []
            )
        ),
        "performance_metrics": ensure_list(
            load_json_file(
                PERFORMANCE_METRICS_FILE,
                []
            )
        ),
        "nlp_actions": ensure_list(
            load_json_file(
                NLP_ACTIONS_FILE,
                []
            )
        ),
        "backup_monitoring": ensure_list(
            load_json_file(
                BACKUP_MONITORING_FILE,
                []
            )
        ),
        "workflow_history": ensure_list(
            load_json_file(
                WORKFLOW_HISTORY_FILE,
                []
            )
        ),
        "ai_recommendations": ensure_list(
            load_json_file(
                AI_RECOMMENDATIONS_FILE,
                []
            )
        )
    }


# =========================================================
# ROW BUILDERS
# =========================================================

def build_summary_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build daily summary row.
    """

    incidents = data.get(
        "incidents",
        []
    )

    actions = data.get(
        "actions",
        []
    )

    pending_approvals = data.get(
        "pending_approvals",
        []
    )

    approval_history = data.get(
        "approval_history",
        []
    )

    healthy_count = 0
    attention_count = 0

    for incident in incidents:

        status = safe_value(
            get_first_available(
                incident,
                [
                    "overall_status",
                    "status"
                ],
                ""
            )
        ).upper()

        if status == "HEALTHY":

            healthy_count += 1

        else:

            attention_count += 1

    approved_count = len(
        [
            approval for approval in approval_history
            if safe_value(
                approval.get(
                    "approval_status",
                    ""
                )
            ).upper() == "APPROVED"
        ]
    )

    rejected_count = len(
        [
            approval for approval in approval_history
            if safe_value(
                approval.get(
                    "approval_status",
                    ""
                )
            ).upper() == "REJECTED"
        ]
    )

    return [
        {
            "Date": current_date(),
            "Total Incidents": len(
                incidents
            ),
            "Healthy Status Count": healthy_count,
            "Attention Status Count": attention_count,
            "Actions Executed": len(
                actions
            ),
            "Approvals Created": len(
                pending_approvals
            ) + len(
                approval_history
            ),
            "Approvals Approved": approved_count,
            "Approvals Rejected": rejected_count
        }
    ]


def build_daily_incident_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Daily Incidents sheet rows.
    """

    rows = []

    for index, incident in enumerate(
        data.get(
            "incidents",
            []
        )
    ):

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        incident,
                        [
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "Incident ID": get_first_available(
                    incident,
                    [
                        "incident_id",
                        "id"
                    ],
                    f"INC-{index + 1:04d}"
                ),
                "Incident Type": get_first_available(
                    incident,
                    [
                        "incident_type",
                        "type"
                    ],
                    "DBA Monitoring"
                ),
                "Severity": get_first_available(
                    incident,
                    [
                        "severity"
                    ],
                    "LOW"
                ),
                "Status": get_first_available(
                    incident,
                    [
                        "overall_status",
                        "status"
                    ],
                    "TRACKED"
                ),
                "Risk Level": get_first_available(
                    incident,
                    [
                        "risk_level",
                        "risk"
                    ],
                    "LOW"
                ),
                "Database Name": get_first_available(
                    incident,
                    [
                        "database_name",
                        "database",
                        "db_name"
                    ],
                    "SQL Server"
                ),
                "Details": get_first_available(
                    incident,
                    [
                        "details",
                        "incident_summary",
                        "summary"
                    ],
                    ""
                ),
                "AI Recommendation": get_first_available(
                    incident,
                    [
                        "ai_recommendation",
                        "recommendation",
                        "ai_analysis",
                        "analysis"
                    ],
                    ""
                )
            }
        )

    return rows


def build_performance_metric_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Performance Metrics sheet rows.
    """

    metrics = data.get(
        "performance_metrics",
        []
    )

    if not metrics:

        metrics = [
            {
                "timestamp": current_timestamp(),
                "cpu_percent": "",
                "memory_percent": "",
                "blocking_sessions": "",
                "deadlocks": "",
                "long_running_queries": "",
                "database_size_gb": "",
                "active_connections": ""
            }
        ]

    rows = []

    for metric in metrics:

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        metric,
                        [
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "CPU %": get_first_available(
                    metric,
                    [
                        "cpu_percent",
                        "cpu",
                        "cpu_usage"
                    ],
                    ""
                ),
                "Memory %": get_first_available(
                    metric,
                    [
                        "memory_percent",
                        "memory",
                        "memory_usage"
                    ],
                    ""
                ),
                "Blocking Sessions": get_first_available(
                    metric,
                    [
                        "blocking_sessions",
                        "blocking_count"
                    ],
                    ""
                ),
                "Deadlocks": get_first_available(
                    metric,
                    [
                        "deadlocks",
                        "deadlock_count"
                    ],
                    ""
                ),
                "Long Running Queries": get_first_available(
                    metric,
                    [
                        "long_running_queries",
                        "long_queries"
                    ],
                    ""
                ),
                "Database Size (GB)": get_first_available(
                    metric,
                    [
                        "database_size_gb",
                        "db_size_gb"
                    ],
                    ""
                ),
                "Active Connections": get_first_available(
                    metric,
                    [
                        "active_connections",
                        "connections"
                    ],
                    ""
                )
            }
        )

    return rows


def build_nlp_action_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build NLP Actions sheet rows.
    """

    rows = []

    for item in data.get(
        "nlp_actions",
        []
    ):

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        item,
                        [
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "User": get_first_available(
                    item,
                    [
                        "user",
                        "performed_by"
                    ],
                    "DBA User"
                ),
                "Natural Language Query": get_first_available(
                    item,
                    [
                        "natural_language_query",
                        "user_query",
                        "query"
                    ],
                    ""
                ),
                "Generated SQL": get_first_available(
                    item,
                    [
                        "generated_sql",
                        "sql_query",
                        "sql"
                    ],
                    ""
                ),
                "Risk Classification": get_first_available(
                    item,
                    [
                        "risk_classification",
                        "risk_level"
                    ],
                    ""
                ),
                "Approval Status": get_first_available(
                    item,
                    [
                        "approval_status"
                    ],
                    ""
                ),
                "Execution Status": get_first_available(
                    item,
                    [
                        "execution_status",
                        "status"
                    ],
                    ""
                )
            }
        )

    return rows


def build_approval_audit_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Approvals Audit sheet rows.
    """

    approvals = []

    approvals.extend(
        data.get(
            "pending_approvals",
            []
        )
    )

    approvals.extend(
        data.get(
            "approval_history",
            []
        )
    )

    rows = []

    for approval in approvals:

        rows.append(
            {
                "Approval ID": get_first_available(
                    approval,
                    [
                        "approval_id",
                        "request_id"
                    ],
                    ""
                ),
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        approval,
                        [
                            "created_at",
                            "timestamp",
                            "decision_at"
                        ]
                    )
                ),
                "Requestor": get_first_available(
                    approval,
                    [
                        "requested_by",
                        "requestor"
                    ],
                    ""
                ),
                "Action Type": get_first_available(
                    approval,
                    [
                        "action_name",
                        "action_type"
                    ],
                    ""
                ),
                "Risk Level": get_first_available(
                    approval,
                    [
                        "risk_level"
                    ],
                    ""
                ),
                "Approval Status": get_first_available(
                    approval,
                    [
                        "approval_status",
                        "status"
                    ],
                    ""
                ),
                "Approver": get_first_available(
                    approval,
                    [
                        "approved_by",
                        "rejected_by",
                        "approver"
                    ],
                    ""
                ),
                "Approval Time": normalize_timestamp(
                    get_first_available(
                        approval,
                        [
                            "decision_at",
                            "approved_at",
                            "rejected_at"
                        ],
                        ""
                    )
                )
            }
        )

    return rows


def build_execution_history_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Execution History sheet rows.
    """

    rows = []

    for index, execution in enumerate(
        data.get(
            "execution_history",
            []
        )
    ):

        rows.append(
            {
                "Execution ID": get_first_available(
                    execution,
                    [
                        "execution_id",
                        "id"
                    ],
                    f"EXE-{index + 1:04d}"
                ),
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        execution,
                        [
                            "executed_at",
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "Action Executed": get_first_available(
                    execution,
                    [
                        "action_executed",
                        "action_name",
                        "action"
                    ],
                    ""
                ),
                "Database": get_first_available(
                    execution,
                    [
                        "database",
                        "database_name",
                        "target_name"
                    ],
                    ""
                ),
                "Execution Status": get_first_available(
                    execution,
                    [
                        "execution_status",
                        "overall_status",
                        "status"
                    ],
                    ""
                ),
                "Duration": get_first_available(
                    execution,
                    [
                        "duration",
                        "duration_seconds"
                    ],
                    ""
                ),
                "Output Summary": get_first_available(
                    execution,
                    [
                        "output_summary",
                        "message",
                        "result"
                    ],
                    ""
                )
            }
        )

    return rows


def build_governance_audit_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Governance Audit sheet rows.
    """

    rows = []

    for event in data.get(
        "governance_audit",
        []
    ):

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        event,
                        [
                            "created_at",
                            "timestamp"
                        ]
                    )
                ),
                "User": get_first_available(
                    event,
                    [
                        "performed_by",
                        "user"
                    ],
                    "System"
                ),
                "Activity Type": get_first_available(
                    event,
                    [
                        "event_type",
                        "activity_type",
                        "event"
                    ],
                    ""
                ),
                "Object Affected": get_first_available(
                    event,
                    [
                        "target_name",
                        "object_affected",
                        "approval_id"
                    ],
                    ""
                ),
                "Result": get_first_available(
                    event,
                    [
                        "status",
                        "result"
                    ],
                    ""
                ),
                "Comments": get_first_available(
                    event,
                    [
                        "message",
                        "comments"
                    ],
                    ""
                )
            }
        )

    return rows


def build_ai_recommendation_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build AI Recommendations sheet rows.
    """

    recommendations = data.get(
        "ai_recommendations",
        []
    )

    if not recommendations:

        for incident in data.get(
            "incidents",
            []
        ):

            recommendation = get_first_available(
                incident,
                [
                    "ai_recommendation",
                    "recommendation",
                    "ai_analysis",
                    "analysis"
                ],
                ""
            )

            if recommendation:

                recommendations.append(
                    {
                        "timestamp": get_first_available(
                            incident,
                            [
                                "timestamp",
                                "created_at"
                            ],
                            current_timestamp()
                        ),
                        "incident_id": get_first_available(
                            incident,
                            [
                                "incident_id"
                            ],
                            ""
                        ),
                        "root_cause": get_first_available(
                            incident,
                            [
                                "root_cause"
                            ],
                            ""
                        ),
                        "recommendation_category": "DBA Recommendation",
                        "recommendation_details": recommendation,
                        "confidence_score": get_first_available(
                            incident,
                            [
                                "confidence_score",
                                "confidence"
                            ],
                            ""
                        )
                    }
                )

    rows = []

    for item in recommendations:

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        item,
                        [
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "Incident ID": get_first_available(
                    item,
                    [
                        "incident_id"
                    ],
                    ""
                ),
                "Root Cause": get_first_available(
                    item,
                    [
                        "root_cause"
                    ],
                    ""
                ),
                "Recommendation Category": get_first_available(
                    item,
                    [
                        "recommendation_category",
                        "category"
                    ],
                    ""
                ),
                "Recommendation Details": get_first_available(
                    item,
                    [
                        "recommendation_details",
                        "recommendation"
                    ],
                    ""
                ),
                "Confidence Score": get_first_available(
                    item,
                    [
                        "confidence_score",
                        "confidence"
                    ],
                    ""
                )
            }
        )

    return rows


def build_backup_monitoring_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Backup Monitoring sheet rows.
    """

    backup_records = data.get(
        "backup_monitoring",
        []
    )

    rows = []

    for item in backup_records:

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        item,
                        [
                            "timestamp",
                            "created_at"
                        ]
                    )
                ),
                "Database": get_first_available(
                    item,
                    [
                        "database",
                        "database_name"
                    ],
                    ""
                ),
                "Last Backup Time": get_first_available(
                    item,
                    [
                        "last_backup_time"
                    ],
                    ""
                ),
                "Backup Type": get_first_available(
                    item,
                    [
                        "backup_type"
                    ],
                    ""
                ),
                "Backup Status": get_first_available(
                    item,
                    [
                        "backup_status",
                        "status"
                    ],
                    ""
                ),
                "Recovery Model": get_first_available(
                    item,
                    [
                        "recovery_model"
                    ],
                    ""
                )
            }
        )

    return rows


def build_agentic_workflow_rows(
    data: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    Build Agentic Workflow History sheet rows.
    """

    workflow_records = data.get(
        "workflow_history",
        []
    )

    rows = []

    for item in workflow_records:

        rows.append(
            {
                "Timestamp": normalize_timestamp(
                    get_first_available(
                        item,
                        [
                            "timestamp",
                            "created_at",
                            "started_at"
                        ]
                    )
                ),
                "Workflow Name": get_first_available(
                    item,
                    [
                        "workflow_name",
                        "name"
                    ],
                    "Agentic DBA Workflow"
                ),
                "Trigger Source": get_first_available(
                    item,
                    [
                        "trigger_source",
                        "source"
                    ],
                    ""
                ),
                "Status": get_first_available(
                    item,
                    [
                        "status",
                        "overall_status"
                    ],
                    ""
                ),
                "Duration": get_first_available(
                    item,
                    [
                        "duration",
                        "duration_seconds"
                    ],
                    ""
                ),
                "Result Summary": get_first_available(
                    item,
                    [
                        "result_summary",
                        "summary",
                        "message"
                    ],
                    ""
                )
            }
        )

    return rows


# =========================================================
# WORKBOOK STYLING
# =========================================================

def style_worksheet(
    worksheet
) -> None:
    """
    Apply professional formatting.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    worksheet.freeze_panes = "A2"

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value:

                max_length = max(
                    max_length,
                    len(
                        str(
                            cell.value
                        )
                    )
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 4,
            60
        )


def add_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: List[str],
    rows: List[Dict[str, Any]]
) -> None:
    """
    Add worksheet with rows.
    """

    worksheet = workbook.create_sheet(
        sheet_name
    )

    worksheet.append(
        headers
    )

    for row in rows:

        worksheet.append(
            [
                safe_value(
                    row.get(
                        header,
                        ""
                    )
                )
                for header in headers
            ]
        )

    style_worksheet(
        worksheet
    )


# =========================================================
# MONTHLY EXCEL GENERATION
# =========================================================

def generate_monthly_dba_excel_report(
    trigger_source: str = "Manual",
    notes: str = "Monthly DBA Excel report generated successfully."
) -> Dict[str, Any]:
    """
    Generate or update monthly DBA Excel report.

    File format:
    DBA_Monthly_Report_YYYY_MM.xlsx
    """

    try:

        ensure_excel_folder()

        report_path = get_monthly_report_file_path()

        platform_data = load_platform_data()

        workbook = Workbook()

        default_sheet = workbook.active

        workbook.remove(
            default_sheet
        )

        sheet_row_builders = {
            "Summary": build_summary_rows(
                platform_data
            ),
            "Daily Incidents": build_daily_incident_rows(
                platform_data
            ),
            "Performance Metrics": build_performance_metric_rows(
                platform_data
            ),
            "NLP Actions": build_nlp_action_rows(
                platform_data
            ),
            "Approvals Audit": build_approval_audit_rows(
                platform_data
            ),
            "Execution History": build_execution_history_rows(
                platform_data
            ),
            "Governance Audit": build_governance_audit_rows(
                platform_data
            ),
            "AI Recommendations": build_ai_recommendation_rows(
                platform_data
            ),
            "Backup Monitoring": build_backup_monitoring_rows(
                platform_data
            ),
            "Agentic Workflow History": build_agentic_workflow_rows(
                platform_data
            )
        }

        for sheet_name, headers in SHEET_DEFINITIONS.items():

            existing_rows = read_existing_rows(
                report_path,
                sheet_name,
                headers
            )

            new_rows = sheet_row_builders.get(
                sheet_name,
                []
            )

            combined_rows = existing_rows + new_rows

            if sheet_name == "Summary":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Date"
                    ]
                )

            elif sheet_name == "Daily Incidents":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Incident ID"
                    ]
                )

            elif sheet_name == "Performance Metrics":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp"
                    ]
                )

            elif sheet_name == "NLP Actions":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Natural Language Query"
                    ]
                )

            elif sheet_name == "Approvals Audit":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Approval ID",
                        "Approval Status"
                    ]
                )

            elif sheet_name == "Execution History":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Execution ID",
                        "Timestamp"
                    ]
                )

            elif sheet_name == "Governance Audit":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Activity Type",
                        "Object Affected"
                    ]
                )

            elif sheet_name == "AI Recommendations":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Incident ID",
                        "Recommendation Details"
                    ]
                )

            elif sheet_name == "Backup Monitoring":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Database",
                        "Backup Type"
                    ]
                )

            elif sheet_name == "Agentic Workflow History":

                combined_rows = unique_rows(
                    combined_rows,
                    [
                        "Timestamp",
                        "Workflow Name",
                        "Trigger Source"
                    ]
                )

            add_sheet(
                workbook,
                sheet_name,
                headers,
                combined_rows
            )

        workbook.save(
            report_path
        )

        return {
            "overall_status": "COMPLETED",
            "report_type": "MONTHLY_DBA_EXCEL_REPORT",
            "report_file": os.path.basename(
                report_path
            ),
            "report_path": report_path.replace(
                "\\",
                "/"
            ),
            "generated_at": current_timestamp(),
            "trigger_source": trigger_source,
            "message": notes
        }

    except Exception as error:

        return {
            "overall_status": "FAILED",
            "report_type": "MONTHLY_DBA_EXCEL_REPORT",
            "message": str(
                error
            )
        }


# =========================================================
# BACKWARD COMPATIBILITY
# =========================================================

def generate_monthly_excel_report(
    *args,
    **kwargs
):
    """
    Backward-compatible alias.
    """

    return generate_monthly_dba_excel_report(
        trigger_source="Backward Compatible Call"
    )


def generate_excel_health_report(
    *args,
    **kwargs
):
    """
    Backward-compatible alias for old report generator.
    """

    return generate_monthly_dba_excel_report(
        trigger_source="Excel Health Report Alias"
    )


# =========================================================
# DIRECT EXECUTION
# =========================================================

if __name__ == "__main__":

    result = generate_monthly_dba_excel_report(
        trigger_source="Manual Script Execution"
    )

    print(
        result
    )