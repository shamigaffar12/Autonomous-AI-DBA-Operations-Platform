# =========================================================
# Excel Health Report Generator
# Autonomous AI DBA Operations Platform
# =========================================================

import json
import os

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# FOLDER CONFIGURATION
# =========================================================

EXCEL_REPORT_FOLDER = "excel_reports"

APPROVAL_FOLDER = "approval_requests"

PENDING_APPROVAL_FILE = os.path.join(
    APPROVAL_FOLDER,
    "pending_approvals.json"
)

APPROVAL_HISTORY_FILE = os.path.join(
    APPROVAL_FOLDER,
    "approval_history.json"
)


# =========================================================
# ENSURE REPORT FOLDER
# =========================================================

def ensure_excel_report_folder():
    """
    Create Excel report output folder if it does not exist.
    """

    os.makedirs(
        EXCEL_REPORT_FOLDER,
        exist_ok=True
    )


# =========================================================
# LOAD JSON FILE
# =========================================================

def load_json_file(
    file_path
):
    """
    Safely load JSON file data.
    """

    try:

        if not os.path.exists(
            file_path
        ):

            return []

        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:

            return json.load(
                file
            )

    except Exception:

        return []


# =========================================================
# STYLE WORKSHEET HEADER
# =========================================================

def style_header(
    worksheet
):
    """
    Apply header style to first row.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


# =========================================================
# AUTO ADJUST COLUMN WIDTH
# =========================================================

def auto_adjust_columns(
    worksheet
):
    """
    Auto adjust column width based on content.
    """

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            try:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(
                            str(
                                cell.value
                            )
                        )
                    )

            except Exception:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 4


# =========================================================
# ADD SUMMARY SHEET
# =========================================================

def add_summary_sheet(
    workbook,
    pending_approvals,
    approval_history
):
    """
    Add executive summary sheet.
    """

    worksheet = workbook.active
    worksheet.title = "Summary"

    summary_data = [
        [
            "Metric",
            "Value"
        ],
        [
            "Report Name",
            "Agentic AI DBA Health Report"
        ],
        [
            "Generated At",
            str(
                datetime.now()
            )
        ],
        [
            "Platform",
            "Autonomous AI DBA Operations Platform"
        ],
        [
            "Database Platform",
            "Microsoft SQL Server"
        ],
        [
            "Monitoring Mode",
            "Live SQL Monitoring + Simulated Azure Adapter"
        ],
        [
            "Pending Approvals",
            len(
                pending_approvals
            )
        ],
        [
            "Approval History Count",
            len(
                approval_history
            )
        ],
        [
            "Azure Monitor",
            "Simulated Adapter Mode"
        ],
        [
            "Azure Automation",
            "Triggered After Approval - Simulated Mode"
        ],
        [
            "Overall Status",
            "Operational MVP"
        ]
    ]

    for row in summary_data:

        worksheet.append(
            row
        )

    style_header(
        worksheet
    )

    auto_adjust_columns(
        worksheet
    )


# =========================================================
# ADD PENDING APPROVALS SHEET
# =========================================================

def add_pending_approvals_sheet(
    workbook,
    pending_approvals
):
    """
    Add pending approval requests sheet.
    """

    worksheet = workbook.create_sheet(
        "Pending Approvals"
    )

    worksheet.append(
        [
            "Approval ID",
            "Action Name",
            "Target Name",
            "Risk Level",
            "Requested By",
            "Status",
            "Reason",
            "Created At"
        ]
    )

    for approval in pending_approvals:

        worksheet.append(
            [
                approval.get(
                    "approval_id"
                ),
                approval.get(
                    "action_name"
                ),
                approval.get(
                    "target_name"
                ),
                approval.get(
                    "risk_level"
                ),
                approval.get(
                    "requested_by"
                ),
                approval.get(
                    "approval_status"
                ),
                approval.get(
                    "reason"
                ),
                approval.get(
                    "created_at"
                )
            ]
        )

    style_header(
        worksheet
    )

    auto_adjust_columns(
        worksheet
    )


# =========================================================
# ADD APPROVAL HISTORY SHEET
# =========================================================

def add_approval_history_sheet(
    workbook,
    approval_history
):
    """
    Add approval history sheet.
    """

    worksheet = workbook.create_sheet(
        "Approval History"
    )

    worksheet.append(
        [
            "Approval ID",
            "Action Name",
            "Target Name",
            "Risk Level",
            "Final Status",
            "Approved By",
            "Rejected By",
            "Decision Reason",
            "Decision At"
        ]
    )

    for approval in approval_history:

        worksheet.append(
            [
                approval.get(
                    "approval_id"
                ),
                approval.get(
                    "action_name"
                ),
                approval.get(
                    "target_name"
                ),
                approval.get(
                    "risk_level"
                ),
                approval.get(
                    "approval_status"
                ),
                approval.get(
                    "approved_by"
                ),
                approval.get(
                    "rejected_by"
                ),
                approval.get(
                    "decision_reason"
                ),
                approval.get(
                    "decision_at"
                )
            ]
        )

    style_header(
        worksheet
    )

    auto_adjust_columns(
        worksheet
    )


# =========================================================
# ADD HEALTH CHECK SHEET
# =========================================================

def add_health_check_sheet(
    workbook
):
    """
    Add health check coverage sheet.
    """

    worksheet = workbook.create_sheet(
        "Health Coverage"
    )

    worksheet.append(
        [
            "Health Area",
            "Current Capability",
            "Remediation Status"
        ]
    )

    health_items = [
        [
            "CPU Usage",
            "Monitoring and recommendation",
            "Recommendation only"
        ],
        [
            "Blocking Sessions",
            "Detection and RCA support",
            "Approval-based extension planned"
        ],
        [
            "Long Running Queries",
            "Detection and tuning recommendation",
            "Recommendation only"
        ],
        [
            "Backup Status",
            "Missing or old backup detection",
            "Approval-based backup automation planned"
        ],
        [
            "Database Space",
            "Space usage monitoring",
            "Storage action planned"
        ],
        [
            "Index Fragmentation",
            "Fragmentation analysis",
            "Maintenance recommendation"
        ],
        [
            "Statistics Health",
            "Stale statistics detection",
            "Update statistics recommendation"
        ],
        [
            "Failed Jobs",
            "Failed job detection",
            "Approval-based restart workflow"
        ]
    ]

    for item in health_items:

        worksheet.append(
            item
        )

    style_header(
        worksheet
    )

    auto_adjust_columns(
        worksheet
    )


# =========================================================
# ADD RECOMMENDATIONS SHEET
# =========================================================

def add_recommendations_sheet(
    workbook
):
    """
    Add production recommendations sheet.
    """

    worksheet = workbook.create_sheet(
        "Recommendations"
    )

    worksheet.append(
        [
            "Area",
            "Recommendation",
            "Priority"
        ]
    )

    recommendations = [
        [
            "Azure Integration",
            "Replace simulated Azure Monitor adapter with real Log Analytics ingestion.",
            "High"
        ],
        [
            "Azure Automation",
            "Trigger real Azure Automation Runbooks after approval.",
            "High"
        ],
        [
            "Approval Workflow",
            "Connect pending approvals with FastAPI UI.",
            "High"
        ],
        [
            "Reporting",
            "Generate Excel-based analytics and health reports.",
            "Completed"
        ],
        [
            "NLP",
            "Add natural language DBA assistant for user queries.",
            "Medium"
        ],
        [
            "Security",
            "Replace hardcoded RBAC role with user-based role mapping.",
            "High"
        ],
        [
            "Notifications",
            "Enable real Email and Microsoft Teams alerts.",
            "Medium"
        ]
    ]

    for recommendation in recommendations:

        worksheet.append(
            recommendation
        )

    style_header(
        worksheet
    )

    auto_adjust_columns(
        worksheet
    )


# =========================================================
# GENERATE EXCEL HEALTH REPORT
# =========================================================

def generate_excel_health_report():
    """
    Generate complete DBA health and governance Excel report.
    """

    try:

        ensure_excel_report_folder()

        pending_approvals = load_json_file(
            PENDING_APPROVAL_FILE
        )

        approval_history = load_json_file(
            APPROVAL_HISTORY_FILE
        )

        workbook = Workbook()

        add_summary_sheet(
            workbook,
            pending_approvals,
            approval_history
        )

        add_health_check_sheet(
            workbook
        )

        add_pending_approvals_sheet(
            workbook,
            pending_approvals
        )

        add_approval_history_sheet(
            workbook,
            approval_history
        )

        add_recommendations_sheet(
            workbook
        )

        report_file_name = (
            "dba_health_analytics_report_"
            + datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + ".xlsx"
        )

        report_file_path = os.path.join(
            EXCEL_REPORT_FOLDER,
            report_file_name
        )

        workbook.save(
            report_file_path
        )

        print("\n========================================")
        print(" Excel Health Report Generated ")
        print("========================================\n")

        print(f"Report Path : {report_file_path}")

        return {
            "overall_status": "COMPLETED",
            "report_type": "EXCEL_HEALTH_ANALYTICS_REPORT",
            "report_path": report_file_path,
            "generated_at": str(
                datetime.now()
            )
        }

    except Exception as error:

        print("\nExcel Report Generation Error:")
        print(error)

        return {
            "overall_status": "ERROR",
            "message": str(
                error
            )
        }


# =========================================================
# TEST EXECUTION
# =========================================================

if __name__ == "__main__":

    result = generate_excel_health_report()

    print(
        result
    )